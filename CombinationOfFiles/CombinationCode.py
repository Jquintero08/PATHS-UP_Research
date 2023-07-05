import pandas as pd
from glob import glob
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import tkinter as tk
from tkinter import filedialog

########################## Change the Master Folder Destination ##########################
##########################################################################################

#desired_location = (r'C:\Users\Jakey\Desktop\Summer_2023_Research\Combination\2021-05-09')


root = tk.Tk()
root.attributes('-alpha', 0.0)  #Makes the extra window fully transparent
desired_location = filedialog.askdirectory()  #Opens the file dialog
root.destroy()  #Gets rid of the main window

##########################################################################################

def main():

   ## Series of code to summarized code from Sergio
   
   SampleRepetition = input('Which repeitition?     ') + '\/'
   SampleRepetition1 = input('Which repeitition?     ') + '\/'

   Spectrum = ["Spectrum"]
   SampleDF = pd.read_csv('Wavenumber-testfile-donotdelete.csv')
   ## Sample file in general location to pull x axis from ##
   SampleDF.columns = ["Wavenumber", "Spectrum", "Standard Deviation"]
   Wavenumber = SampleDF.loc[0:,'Wavenumber']
   print(Wavenumber)
   NewDF = pd.DataFrame()

   OS_finalloc = desired_location + SampleRepetition
   print(OS_finalloc)
   sorted_OS_listdir = sorted(os.listdir(OS_finalloc))

   for file1 in sorted_OS_listdir :
       df_file = pd.read_csv(OS_finalloc + file1)
       df_file.columns = ["Integration ", "S1",]
       print(df_file.columns)
       df_fileloc = df_file.loc[15:,['S1']]
       df_fileloc.reset_index(drop=True, inplace=True)
       print(df_fileloc)
       NewDF = pd.concat([df_fileloc, NewDF], axis=1)
       NewDF.reset_index(drop=True, inplace=True)
       #print(NewDF)
   NewDF2 = pd.concat([Wavenumber, NewDF], axis=1)
   #print(NewDF2)
   NewDF2.reset_index(drop=True, inplace=True)


   Outfile_pathq = input('What folder name would you like?       ')
   ParentDir = (r'C:\Users\Angela\PycharmProjects\CompilingSERS\/')
   Outfile_path = (r'C:\Users\Angela\PycharmProjects\CompilingSERS\/') + Outfile_pathq
   #print(Outfile_path)
   Outdir = os.mkdir(Outfile_path)
   #print(Outdir)

   OutFileDF = input('What will you name your fle for your set of samples -MatlabProcessing  ')
   OutFileDF2 = input('What will you name your fle for your set of samples + with header/wavenumber   ')
   NewDFOut = (OutFileDF + '.csv')
   NewDFOut2 = (OutFileDF2 + '.csv')


   NewDF.to_csv(os.path.join(Outfile_path, NewDFOut), header = False, index = False)
   NewDF2.to_csv(os.path.join(Outfile_path, NewDFOut2), header = True)
   restart = input('Would you like to compile more data?      ')
   YesList = ["Yes", "y", "Y", "yes"]
   if restart in YesList :
       main()
   else:
       quit()
       
       
     
def combine_files(base_dir):
    #Create a new Workbook
    wb = Workbook()

    #Removes default worksheet and produced blank workbook
    default_sheet = wb.active

    #Go through all files and directories in the Master Folder
    for root, dirs, files in os.walk(base_dir):
        for dir in dirs:
            #Produces list of all .csv files
            csv_files = [file for file in os.listdir(os.path.join(root, dir)) if file.endswith('.csv')]

            # Skip this directory if it does not contain any csv files
            if not csv_files:
                continue

            #Creates new worksheet for the directory with the directory name
            ws = wb.create_sheet(title=dir)

            #Go through all .csv files in directory and combine them
            dataframes = []
            for i, file in enumerate(csv_files):
                df = pd.read_csv(os.path.join(root, dir, file), skiprows=16, header=None)
                df.columns = df.iloc[0]
                df = df[1:]
                
                #If its not the first file in the directory, only add the second column
                if i > 0:
                    df = df.iloc[:, 1:2]  #Adjust here if column numbering doesn't start from 0
                dataframes.append(df)

            all_data = pd.concat(dataframes, axis=1)

            #Add data to the singular worksheet
            for row in dataframe_to_rows(all_data, index=False, header=True):
                ws.append(row)

    #If nothing created, delete default sheet
    if len(wb.sheetnames) > 1:
        wb.remove(default_sheet)

    #Save workbook to this path
    output_file_name = input("Enter the output file name (without extension): ")
    output_file_name = output_file_name + '.xlsx'
    
    
    ################################ Change the output file destination #########################################
    #############################################################################################################
    
    #output_dir = (r'C:\Users\Jakey\Desktop\Summer_2023_Research\Combination\2021-05-09') #Uncomment if you don't want to select file output & comment root = tk.Tk() to output_file_path = ...
    #output_file_path = os.path.join(output_dir, output_file_name)
    
    root = tk.Tk()
    root.attributes('-alpha', 0.0)  #Makes the extra window transparent
    output_dir = filedialog.askdirectory()  #Opens the file dialog
    root.destroy()  #Gets rid of the main window
    output_file_path = os.path.join(output_dir, output_file_name) #Join the output location & output file name
    
    #############################################################################################################
    
    wb.save(output_file_path)
 
combine_files(desired_location)
main()
