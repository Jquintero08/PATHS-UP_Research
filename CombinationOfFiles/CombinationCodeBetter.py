import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import tkinter as tk
from tkinter import filedialog

########################## Change the Master Folder Destination ##########################
##########################################################################################

#desired_location = (r'C:\Users\Jakey\Desktop\Summer_2023_Research\Combination\2021-05-09')


root = tk.Tk()
root.attributes('-alpha', 0.0)  #Makes the extra window fully transparent
desiredLocation = filedialog.askdirectory()  #Opens the file dialogue
root.destroy()  #Gets rid of the main window

##########################################################################################
       
     
def main(base_dir):
    #Create a new Workbook
    wb = Workbook()

    #Removes default worksheet and produced blank workbook
    default_sheet = wb.active

    #Go through all files and directories in the Master Folder
    for root, dirs, files in os.walk(base_dir):
        for dir in dirs:
            #Produces list of all .csv files
            csv_files = [file for file in os.listdir(os.path.join(root, dir)) if file.endswith('.csv')]

            #Skip directory if it does not contain any csv files
            if not csv_files:
                continue

            #Creates new worksheet for the directory with the directory name
            ws = wb.create_sheet(title=dir)
            
            #isFirstFile = True #Comment out 

            #Go through all .csv files in directory and combine them
            dataframes = []
            for i, file in enumerate(csv_files):
                df = pd.read_csv(os.path.join(root, dir, file), skiprows=17, header=None)
                df.columns = df.iloc[0]
                df = df[1:]
                
                #if isFirstFile: #Comment out
                #    df = df.iloc[:, :2] #Comment out
                #    isFirstFile = False #Comment out
                #else: #Comment out
                df = df.iloc[:, 1:2]  #Adjust here if column numbering doesn't start from 0 (Deindent)
                   
                dataframes.append(df)

            all_data = pd.concat(dataframes, axis=1)

            #Add data to the singular worksheet
            for row in dataframe_to_rows(all_data, index=False, header=True):
                ws.append(row)

    #If nothing created, delete default sheet
    if len(wb.sheetnames) > 1:
        wb.remove(default_sheet)

    #Save workbook to this path
    #output_file_name = input("Enter the output file name (without extension): ") #For input in terminal
    #output_file_name = output_file_name + '.xlsx'
    
    
    ################################ Change the output file destination #########################################
    #############################################################################################################
    
    #output_dir = (r'C:\Users\Jakey\Desktop\Summer_2023_Research\Combination\2021-05-09') #Uncomment if you don't want to select file output & comment root = tk.Tk() to output_file_path = ...
    #output_file_path = os.path.join(output_dir, output_file_name)
    
    root = tk.Tk()
    root.attributes('-alpha', 0.0)  #Makes the extra window transparent
    #output_dir = filedialog.askdirectory()  #Opens the file dialog (For input in terminal)
    outputFile = filedialog.asksaveasfilename(filetypes=[('Excel file', '*.xlsx')])
    root.destroy()  #Gets rid of the main window
    #output_file_path = os.path.join(output_dir, output_file_name) #Join the output location & output file name
    #Input in terminal code rather than pop-up ^
    
    #############################################################################################################
    
    wb.save(outputFile)
    
    print("\nDone!")
 
main(desiredLocation)